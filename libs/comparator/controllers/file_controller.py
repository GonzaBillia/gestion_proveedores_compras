import logging
import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils import get_column_letter

def setup_logger():
    """Setup logger for the file controller."""
    logging.basicConfig(
        level=logging.DEBUG,
        format='%(asctime)s - %(levelname)s - %(message)s',
        handlers=[
            logging.FileHandler("file_controller.log"),
            logging.StreamHandler()
        ]
    )

setup_logger()

def open_file():
    """Open a file by prompting the user for its path."""
    try:
        file_path = input("Enter the file path to open: ")
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"The file {file_path} does not exist.")
        logging.info(f"File opened successfully: {file_path}")
        return file_path
    except FileNotFoundError as e:
        logging.error(f"Error opening file: {e}")
        return None
    except Exception as e:
        logging.error(f"Unexpected error opening file: {e}")
        return None

def save_file():
    """Prompt the user for a destination path to save a file."""
    try:
        destination_path = input("Enter the destination path to save the file: ")
        directory = os.path.dirname(destination_path)
        if directory and not os.path.exists(directory):
            raise FileNotFoundError(f"The directory {directory} does not exist.")
        logging.info(f"File destination selected: {destination_path}")
        return destination_path
    except FileNotFoundError as e:
        logging.error(f"Error selecting file destination: {e}")
        return None
    except Exception as e:
        logging.error(f"Unexpected error selecting file destination: {e}")
        return None

def read_file(file_path):
    """Read a file and return its content as a DataFrame if applicable."""
    try:
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"The file {file_path} does not exist.")
        
        file_extension = os.path.splitext(file_path)[1].lower()
        if file_extension == ".csv":
            df = pd.read_csv(file_path)
            logging.info(f"CSV file read successfully: {file_path}")
            return df
        elif file_extension == ".xlsx":
            df = pd.read_excel(file_path)
            logging.info(f"Excel file read successfully: {file_path}")
            return df
        else:
            with open(file_path, 'r') as file:
                content = file.read()
            logging.info(f"Text file read successfully: {file_path}")
            return content
    except FileNotFoundError as e:
        logging.error(f"Error reading file: {e}")
        return None
    except pd.errors.EmptyDataError as e:
        logging.error(f"The file is empty or corrupted: {file_path} - {e}")
        return None
    except Exception as e:
        logging.error(f"Unexpected error reading file: {e}")
        return None
    
def normalize_columns(df):
    """
    Normaliza las columnas de un DataFrame:
    - Elimina espacios iniciales y finales.
    - Convierte los nombres de las columnas a minúsculas.

    :param df: DataFrame a normalizar.
    :return: DataFrame con columnas normalizadas.
    """
    df.columns = df.columns.str.strip().str.lower()
    return df

def export_file(dataframe, filename, file_format='excel', sheet_name='Sheet1'):
    """
    Exporta un DataFrame a un archivo en el formato especificado.

    Args:
        dataframe (pd.DataFrame): El DataFrame a exportar.
        filename (str): Nombre del archivo a exportar (sin extensión).
        file_format (str): Formato del archivo ('excel', 'csv'). Default: 'excel'.
        sheet_name (str): Nombre de la hoja (solo aplica para Excel). Default: 'Sheet1'.
    """
    if file_format.lower() == 'excel':
        # Exportar como archivo Excel
        full_filename = f"{filename}.xlsx"
        dataframe.to_excel(full_filename, index=False, sheet_name=sheet_name)
    elif file_format.lower() == 'csv':
        # Exportar como archivo CSV
        full_filename = f"{filename}.csv"
        dataframe.to_csv(full_filename, index=False)
    else:
        raise ValueError("Formato no soportado. Usa 'excel' o 'csv'.")

    print(f"Archivo exportado: {full_filename}")

import pandas as pd

def export_file_to_excel(dataframes_with_names, filename):
    path = 'C:\\Users\\Administrador\\Documents\\Gonzalo\\archivos\\lista proveedores\\comparados\\'

    full_path = path + filename
    """
    Exporta múltiples DataFrames a un archivo Excel con hojas separadas.

    Args:
        dataframes_with_names (list of tuples): Lista de tuplas donde cada una contiene un DataFrame y el nombre de la hoja.
            Ejemplo: [(df1, 'Sheet1'), (df2, 'Sheet2')]
        filename (str): Nombre del archivo Excel a exportar. Ejemplo: 'results.xlsx'.
    """

    if not dataframes_with_names:
        print("No se proporcionaron DataFrames para exportar.")
        return

    # Exportar a Excel
    with pd.ExcelWriter(full_path, engine='openpyxl') as writer:
        for dataframe, sheet_name in dataframes_with_names:
            # Verificar que el DataFrame no esté vacío y sea válido
            if dataframe is not None and not dataframe.empty:
                dataframe.to_excel(writer, index=False, sheet_name=sheet_name[:31])  # Limitar el nombre a 31 caracteres
            else:
                print(f"El DataFrame asociado a la hoja '{sheet_name}' está vacío o no existe.")

    # Cargar el archivo con openpyxl para aplicar estilos
    workbook = load_workbook(full_path)
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        apply_styles_to_sheet(sheet)
    workbook.save(full_path)

    print(f"Archivo exportado exitosamente: {filename}")

    return full_path

# Función para aplicar estilos a cada hoja
def apply_styles_to_sheet(sheet):
    # Añadir estilos a las celdas
    for row in sheet.iter_rows():
        for cell in row:
            # Añadir bordes
            thin_border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin")
            )
            cell.border = thin_border
            # Centrar el texto
            cell.alignment = Alignment(horizontal="center", vertical="center")
            # Aplicar estilo especial a la primera fila (cabecera)
            if cell.row == 1:
                cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Color amarillo
                cell.font = Font(bold=True)  # Texto en negrita
    
    # Ajustar el ancho de las columnas
    for col_idx in range(1, sheet.max_column + 1):
        column_letter = get_column_letter(col_idx)
        max_length = 0
        for cell in sheet[column_letter]:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        sheet.column_dimensions[column_letter].width = max_length + 2  # Ajustar con un margen extra

def format_costs_excel(input_path, quantio_col="C", provider_col="D"):
    """
    Carga un archivo Excel, aplica formato condicional para resaltar diferencias en precios entre dos columnas y guarda el archivo.

    Args:
        input_path (str): Ruta del archivo Excel de entrada.
        output_path (str): Ruta donde se guardará el archivo Excel modificado.
        quantio_col (str): Letra de la columna con los precios de Quantio.
        provider_col (str): Letra de la columna con los precios del Proveedor.
    """
    # Cargar el archivo Excel existente
    wb = load_workbook(input_path)
    ws = wb.active

    # Aplicar formato condicional
    red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")  # Rojo
    green_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")  # Verde

    # Fórmula para cuando Costo Quantio > Costo Proveedor
    formula_red = f"${quantio_col}2>${provider_col}2"
    rule_red = FormulaRule(formula=[formula_red], stopIfTrue=True, fill=red_fill)

    # Fórmula para cuando Costo Quantio <= Costo Proveedor
    formula_green = f"${quantio_col}2<={provider_col}2"
    rule_green = FormulaRule(formula=[formula_green], stopIfTrue=True, fill=green_fill)

    # Aplicar las reglas a la columna Costo Quantio
    rango = f"{quantio_col}2:{quantio_col}{ws.max_row}"
    ws.conditional_formatting.add(rango, rule_red)
    ws.conditional_formatting.add(rango, rule_green)

    # Guardar el archivo modificado
    wb.save(input_path)
