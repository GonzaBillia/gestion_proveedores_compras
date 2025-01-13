import os
import json
import pandas as pd
from PyQt5.QtWidgets import QFileDialog, QMessageBox, QInputDialog
from controllers.preferences_controller import PreferencesController
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook

class FileController:
    def __init__(self, data_processor):
        """
        Inicializa el gestor de archivos.
        """
        self.preferences_controller = PreferencesController()
        self.data_processor = data_processor
        self.directories_file = self.preferences_controller.directories_file
        self.preferences = self.load_preferences()
        self.file_names = []  # Lista de rutas de archivos cargados
        self.sheet_selection = {}  # Diccionario con la selección de hojas por archivo
        self._current_index = 0  # Índice para iterar sobre los archivos cargados

    def load_preferences(self):
        """
        Carga las preferencias desde el archivo JSON.
        :return: Un diccionario con las preferencias.
        """
        if os.path.exists(self.directories_file):
            with open(self.directories_file, "r") as file:
                preferences = json.load(file)
            return preferences
        return {}
    
    def get_preference_path(self, directory_key):
        """
        Obtiene la ruta de preferencia y si debe preguntar dónde guardar.
        :param directory_key: Clave del directorio en las preferencias.
        :return: Tuple con la ruta y el booleano de preguntar.
        """
        directory = self.preferences.get("directories", {}).get(directory_key, {})
        path = directory.get("path", None)
        ask = directory.get("ask", True)
        return path, ask

    def save_combined_file(self, directory_key):
        """
        Guarda el archivo combinado en la ubicación especificada o según las preferencias del usuario.
        :param directory_key: Clave del directorio para buscar en las preferencias.
        """
        try:
            preferred_path, ask_where_to_save = self.get_preference_path(directory_key)

            if ask_where_to_save or not preferred_path:
                # Si la preferencia es preguntar o no hay una ruta de preferencia, abrir diálogo
                output_file, _ = QFileDialog.getSaveFileName(None, "Guardar archivo combinado", "", "Excel files (*.xlsx)")
            else:
                # Si hay una ruta de preferencia configurada, usarla
                filename = self.ask_filename_dialog()
                output_file = os.path.join(preferred_path, f"{filename}.xlsx")

            if output_file:
                self.data_processor.save_combined_file(output_file)
                
                # Cargar el archivo con openpyxl para aplicar estilos
                workbook = load_workbook(output_file)
                for sheet_name in workbook.sheetnames:
                    sheet = workbook[sheet_name]
                    apply_styles_to_sheet(sheet)
                workbook.save(output_file)

                QMessageBox.information(None, "Éxito", f"Archivo combinado guardado en: {output_file}")
                os.startfile(output_file)
            else:
                QMessageBox.warning(None, "Advertencia", "No se seleccionó ninguna ubicación para guardar el archivo.")
        except Exception as e:
            QMessageBox.critical(None, "Error", f"Error al guardar el archivo: {e}")


    def load_files(self, file_paths):
        """
        Carga los archivos seleccionados por el usuario.
        :param file_paths: Lista de rutas de archivos Excel.
        """
        self.file_names = file_paths
        self.sheet_selection = {file: None for file in self.file_names}
        self._current_index = 0  # Reinicia el índice para la iteración

    def set_sheet_selection(self, file, sheets):
        """
        Almacena las hojas seleccionadas para un archivo.
        :param file: Ruta del archivo.
        :param sheets: Lista de hojas seleccionadas.
        """
        if file in self.file_names:
            self.sheet_selection[file] = sheets
        else:
            raise ValueError(f"El archivo {file} no está cargado.")

    def get_sheet_names(self, file):
        """
        Obtiene los nombres de las hojas de un archivo Excel.
        :param file: Ruta del archivo Excel.
        :return: Lista de nombres de hojas.
        """
        try:
            return pd.ExcelFile(file).sheet_names
        except Exception as e:
            print(f"Error al obtener hojas de {file}: {e}")
            return []

    def read_excel(self, file, sheet_name=None, header_row=0):
        """
        Lee un archivo Excel con la hoja y encabezado especificados.
        :param file: Ruta del archivo Excel.
        :param sheet_name: Nombre de la hoja (opcional).
        :param header_row: Fila de encabezado (por defecto: 0).
        :return: DataFrame con los datos leídos.
        """
        try:
            return pd.read_excel(file, sheet_name=sheet_name, header=header_row, dtype=str)
        except Exception as e:
            print(f"Error al leer el archivo {file} (hoja: {sheet_name}): {e}")
            return pd.DataFrame()  # Devuelve un DataFrame vacío en caso de error

    def read_selected_sheets(self, header_row=0):
        """
        Lee las hojas seleccionadas para todos los archivos cargados.
        :param header_row: Fila de encabezado (por defecto: 0).
        :return: Diccionario con DataFrames por archivo y hoja.
        """
        data = {}
        for file, sheets in self.sheet_selection.items():
            if sheets is None:  # Si no hay hojas seleccionadas, leer todo el archivo
                data[file] = {None: self.read_excel(file, header_row=header_row)}
            else:
                data[file] = {
                    sheet: self.read_excel(file, sheet_name=sheet, header_row=header_row) for sheet in sheets
                }
        return data

    def validate_files(self):
        """
        Valida los archivos cargados para asegurarse de que sean válidos y legibles.
        :return: Diccionario con el estado de validación de cada archivo.
        """
        validation_results = {}
        for file in self.file_names:
            try:
                pd.ExcelFile(file)  # Intenta abrir el archivo para validar
                validation_results[file] = "Válido"
            except Exception as e:
                validation_results[file] = f"Inválido: {e}"
        return validation_results

    def combine_dataframes(self, data, selected_columns=None):
        """
        Combina múltiples DataFrames seleccionados en un único DataFrame.
        :param data: Diccionario con DataFrames por archivo y hoja.
        :param selected_columns: Lista de columnas a incluir en el DataFrame combinado (opcional).
        :return: DataFrame combinado.
        """
        combined_df = pd.DataFrame()
        for file, sheets in data.items():
            for sheet, df in sheets.items():
                if not df.empty:
                    if selected_columns:
                        missing_columns = [col for col in selected_columns if col not in df.columns]
                        if missing_columns:
                            print(f"Advertencia: Faltan columnas {missing_columns} en {file} - {sheet}")
                            continue
                        df = df[selected_columns]  # Filtrar columnas si se especifican
                    combined_df = pd.concat([combined_df, df], ignore_index=True)
        return combined_df

    def get_next_file(self):
        """
        Iterador que devuelve el siguiente archivo y hoja seleccionada.
        """
        while self._current_index < len(self.file_names):
            file = self.file_names[self._current_index]
            sheets = self.sheet_selection.get(file, [None])  # Si no hay hojas seleccionadas, procesar todo el archivo
            for sheet in sheets:
                yield file, sheet
            self._current_index += 1
        raise StopIteration("No quedan más archivos para procesar.")
    
    def ask_filename_dialog(self):
        """
        Muestra un cuadro de diálogo para pedir al usuario el nombre del archivo.
        :return: El nombre del archivo como una cadena de texto.
        """
        filename, ok = QInputDialog.getText(None, "Guardar archivo", "Ingrese el nombre del archivo:")
        if ok and filename:
            return filename
        return None

# Función para aplicar estilos a cada hoja
def apply_styles_to_sheet(sheet):
    # Añadir estilos a las celdas
    for row in sheet.iter_rows():
        for cell in row:
            # Añadir bordes
            thin_border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin")
            )
            cell.border = thin_border
            # Centrar el texto
            cell.alignment = Alignment(horizontal="center", vertical="center")
            # Aplicar estilo especial a la primera fila (cabecera)
            if cell.row == 1:
                cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Color amarillo
                cell.font = Font(bold=True)  # Texto en negrita
    
    # Ajustar el ancho de las columnas
    for col_idx in range(1, sheet.max_column + 1):
        column_letter = get_column_letter(col_idx)
        max_length = 0
        for cell in sheet[column_letter]:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        sheet.column_dimensions[column_letter].width = max_length + 2  # Ajustar con un margen extra